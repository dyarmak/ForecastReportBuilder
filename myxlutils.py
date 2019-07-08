import openpyxl
import xlrd
from openpyxl import Workbook
from datetime import datetime

# Create a list of due columns for NEXT Year
def next_year_due_columns():
    """Return a list of columns for the "Due" field for NEXT year 
    """
    today = datetime.now().date()
    curryear = today.year
    next_year = datetime(curryear+1, 1, 1)
    dueCols = []

    for p in range(next_year.month, 13):
        dtString = str(p) + str(next_year.year)
        dt = datetime.strptime(dtString, '%m%Y').date()
        futureString = dt.strftime('%m') + "-" + dt.strftime('%Y')
        dueCols.append(futureString) 

    return dueCols

def due_columns():
    """
    Returns a list [] of columns for the "Due" Field
    this would not work on past data...
    It should be modified to somehow use data from the pandas DF.
    """
    today = datetime.now().date()
    curryear = today.strftime("%Y")
    dueCols = ["Def-1"]

    # Act month
    for p in range(1, today.month+1):
        dtString = str(p) + str(curryear)
        dt = datetime.strptime(dtString, '%m%Y').date()
        actString = "Act"+ dt.strftime('%m') + "-" + dt.strftime('%Y')
        dueCols.append(actString)

    # Overdue, completed, Review
    dueCols.append("Overdue")
    dueCols.append("Complete")
    dueCols.append("Review")

# Future months
    for p in range(today.month, 13):
        dtString = str(p) + str(curryear)
        dt = datetime.strptime(dtString, '%m%Y').date()
        futureString = dt.strftime('%m') + "-" + dt.strftime('%Y')
        dueCols.append(futureString) 

    # Should check the generated list against the a list of unique due values from Pandas?

    return dueCols


def format_dollar_values(sheetVar, colNamesDict, formatString, *args):
    """
    Apply currency formatting to rows in dictionary
    sheetVar = variable containing sheet
    colNamesDict = dictionary of all column names in sheet
    formatString = Should be "mm-dd-yy" for it to work
    *args = the columnNames we want to apply date formatting to
    """
    
    for r in range(2, sheetVar.max_row+1):
        for columnName in args:
            sheetVar.cell(row=r, column=colNamesDict[columnName]).number_format = formatString


def cvt_xls_to_xlsx(src_file_path, dst_file_path):
    """
    Takes .xls file as input and output .xlsx file with cell contents copied

    OpenPyXL can only work with .xlsx files (Excel 2007 and newer)

    """
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index in range(0,len(sheet_names)):
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)

    book_xlsx.save(dst_file_path)


def get_column_names_and_index(sheetVar, emptyDict):
    """
    Fill an empty dictionary with the column names
    sheetVar = variable containing sheet
    emptyDict = an empty dictionary to hold column names
    """
    for x in range(1,(sheetVar.max_column+1)):
        emptyDict[str(sheetVar.cell(row=1, column=x).value)] = x
    
    # print("Values added to dictionary")


def format_date_rows(sheetVar, colNamesDict, formatString, *args):
    """
    Apply date formatting to rows in dictionary
    sheetVar = variable containing sheet
    colNamesDict = dictionary of all column names in sheet
    formatString = Should be "mm-dd-yy" for it to work
    *args = the columnNames we want to apply date formatting to
    """
    
    for r in range(2, sheetVar.max_row+1):
        for columnName in args:
            sheetVar.cell(row=r, column=colNamesDict[columnName]).number_format = formatString

    # print("date rows formatted")

def save_and_reopen(wbVar, sheetVar, tempFileName):
    """
    Save a temp file and reopen it to apply the date formatting
    wbVar = current woorkbook variable
    sheetVar = current sheet variable
    tempFileName = filename String
    """
    if(tempFileName.endswith('.xlsx')):
        wbVar.save(tempFileName)
        print(tempFileName + " was created")
        wbVar.close()
        wbVar = openpyxl.load_workbook(tempFileName)
        sheetVar = wbVar.active
        print("Closed and Re-Opened workbook so dates play nicely")
    else:
        print(tempFileName + " must be .xlsx")
        