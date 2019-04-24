#************************************************************************
# This script performs the excel manipulations on the qry_Credits data
#************************************************************************

# Currently broken
# Need to add 'Type' Column in correct loaction

import os
import glob
import datetime
import time
import xlrd
from openpyxl import Workbook
import openpyxl
from myxlutils import save_and_reopen, format_date_rows, get_column_names_and_index
from excelFNames import creditFName, logMe

startTimer = time.time()

# ------------- Load ---------------
wbCred = openpyxl.load_workbook(creditFName)
sCred = wbCred.active


# ----------- BEGIN Find Column Indexes -------------------
print("Loading Credits column names")
#Create an empty dictionary
creditsDict = {}
# fill the dictionary with:
# Key = column names
# Value = column index
get_column_names_and_index(sCred, creditsDict)


# Append "CR" to SubProjectID

for r in range(2, sCred.max_row+1):
    sCred.cell(row=r, column=creditsDict["SubProjectID"]).value = str(sCred.cell(row=r, column=creditsDict["SubProjectID"]).value) +"CR"
print("Added 'CR' after SubProjectIDs")



# ------------ SAVE Output Excel File -------------

wbCred.save(creditFName)
print("Manipulated Credits query saved as " + creditFName)
wbCred.close()


# -------------- Timer -----------------------
endTimer = time.time()
creditsTime = endTimer-startTimer
print("Execution on Credits took: " + str(creditsTime) + "seconds to execute\n")
