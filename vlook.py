#*************************************************************************************
# This script will load the vlook.xlsx file and set 'Type' column values accordingly
#*************************************************************************************

import os
import openpyxl
import pandas as pd
from excelFNames import combinedFName
from myxlutils import get_column_names_and_index
from openpyxl.utils import get_column_letter
from paths import startPath, savePath
# **************************************************************************************************************
# ------------------- Load vlookup table from a txt or excel file ------------------------
# **************************************************************************************************************
os.chdir(startPath) # vlook.csv is in the start path
vlook = pd.read_csv('vlook.csv', header=None, index_col=0, squeeze=True).to_dict()
os.chdir(savePath)

wbCombined = openpyxl.load_workbook(combinedFName)
sCombined = wbCombined.active
combinedDict = {}
get_column_names_and_index(sCombined, combinedDict)
column = get_column_letter(combinedDict["SubProjectTypeName"])

for r in range(2, sCombined.max_row+1):
        if(sCombined.cell(row=r, column = combinedDict["MasterProjectName"]).value == "Software"):
                sCombined.cell(row=r, column = combinedDict["Type"]).value = "Software" 

        elif(sCombined.cell(row=r, column = combinedDict["MasterProjectName"]).value == "HSE"):
                sCombined.cell(row=r, column = combinedDict["Type"]).value = "HSE" 

        elif(sCombined.cell(row=r, column = combinedDict["MasterProjectName"]).value == "CRB"):
                sCombined.cell(row=r, column = combinedDict["Type"]).value = "CRB" 

        else:
                lookup = sCombined.cell(row=r, column = combinedDict["SubProjectTypeName"]).value
                sCombined.cell(row=r, column= combinedDict["Type"]).value = vlook.get(lookup, None)


wbCombined.save(combinedFName)
wbCombined.close()

print("Value of type column set as per values in 'vlook.csv' table\n")