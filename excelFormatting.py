import os
import re
import openpyxl
from excelFNames import combinedFName, unformattedFName, outputFName
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter
from myxlutils import format_date_rows, get_column_names_and_index, format_dollar_values
from paths import savePath # Need to be in the py_output folder

# Open in Openpyxl and apply some cell formatting
# Load workbook
wbReport = openpyxl.load_workbook(unformattedFName)

# Formatting vars
blueFill = PatternFill(patternType="solid", fgColor="91B3D7")
currentFill = PatternFill(patternType="solid", fgColor="E6B8B7")
redBorder = Side(border_style="medium", color = "FF0000")
boldFont = Font(bold=True)
leftAlign = Alignment(horizontal="left", vertical="top")

# *****************************************************************
# ************************ Details Tab ****************************
# *****************************************************************

# Load details sheet
sDetail = wbReport["Details"]

detailDict = {}
get_column_names_and_index(sDetail, detailDict)

# Freeze and left-align top row
sDetail.freeze_panes = "A2"

for x in range(1,(sDetail.max_column+1)):
    sDetail.cell(row=1, column=x).alignment = leftAlign

# Format dates to mm-dd-yyyy
format_date_rows(sDetail, detailDict, "mm-dd-yy", "Due Date", "InvoiceDateSent", "OriginalDueDate")

# Format currency cols to ()
format_dollar_values(sDetail,detailDict, '"$"#,##0_);[Red]("$"#,##0)', "Forecast", "Quoted", "OriginalForecast", "Budget", "SubTotal")

# Format Column Widths
sDetail.column_dimensions[get_column_letter(detailDict['SubProjectID'])].width = 13
sDetail.column_dimensions[get_column_letter(detailDict['ProjectManager'])].width = 15.4
sDetail.column_dimensions[get_column_letter(detailDict['ClientName'])].width = 17
sDetail.column_dimensions[get_column_letter(detailDict['MasterProjectName'])].width = 19
sDetail.column_dimensions[get_column_letter(detailDict['ProjectName'])].width = 13
sDetail.column_dimensions[get_column_letter(detailDict['Type'])].width = 15
sDetail.column_dimensions[get_column_letter(detailDict['SubProjectTypeName'])].width = 20
sDetail.column_dimensions[get_column_letter(detailDict['SubProjectName'])].width = 17
sDetail.column_dimensions[get_column_letter(detailDict['Forecast'])].width = 10
sDetail.column_dimensions[get_column_letter(detailDict['Due Date'])].width = 15.2
sDetail.column_dimensions[get_column_letter(detailDict['Due'])].width = 13
sDetail.column_dimensions[get_column_letter(detailDict['SubProjectStatus'])].width = 20
sDetail.column_dimensions[get_column_letter(detailDict['Quoted'])].width = 10
sDetail.column_dimensions[get_column_letter(detailDict['OriginalForecast'])].width = 14.3
sDetail.column_dimensions[get_column_letter(detailDict['OriginalDueDate'])].width = 15.3
sDetail.column_dimensions[get_column_letter(detailDict['Budget'])].width = 9
sDetail.column_dimensions[get_column_letter(detailDict['InvoiceDateSent'])].width = 15.2


yellowFill = PatternFill(patternType="solid", fgColor="FFFF00" )
for r in range(2, sDetail.max_row+1):
        # IF InvoiceDateSent != None, Forecast GOT None But formatting doesn't copy over. 
        if(sDetail.cell(row=r, column=detailDict["InvoiceDateSent"]).value != None): # May need to test for == 2018 or == 2019 instead
                sDetail.cell(row=r, column=detailDict["Forecast"]).fill = yellowFill 


# *****************************************************************
# ************************ ByType Tab *****************************
# *****************************************************************

sByType = wbReport["ByType"]

# This really is the profit center type.
# In QBO this is the "Class" field so lets rename it that.
sByType.cell(row=1, column = 1).value = "Class"

typeDict = {}
get_column_names_and_index(sByType, typeDict)

# Freeze and left-align top row
sByType.freeze_panes = "A2"

# Format currency cols to ()
for r in range(2, sByType.max_row+1):
    for c in range(2, sByType.max_column+1):
        sByType.cell(row=r, column=c).number_format = '"$"#,##0_);[Red]("$"#,##0)'

# Format Column Widths
for col in range(2, sByType.max_column):
    sByType.column_dimensions[get_column_letter(col)].width = 12
sByType.column_dimensions[get_column_letter(1)].width = 20
sByType.column_dimensions[get_column_letter(sByType.max_column)].width = 14

# Color Total Rows and Cols blue and make bold and left align
for c in range(1, sByType.max_column):
    sByType.cell(row = sByType.max_row, column = c).fill = blueFill
    sByType.cell(row = sByType.max_row, column = c).font = boldFont
for r in range(2, sByType.max_row+1):
    sByType.cell(row = r, column = typeDict["Total"]).fill = blueFill
    sByType.cell(row = r, column = typeDict["Total"]).font = boldFont
    sByType.cell(row = r, column = 1).alignment = leftAlign

# Color Overdue, Complete, Review, Review+1 #E6B8B7
sByType.cell(row = 1, column = typeDict["Overdue"]-1).fill = currentFill
sByType.cell(row = 1, column = typeDict["Overdue"]).fill = currentFill
sByType.cell(row = 1, column = typeDict["Complete"]).fill = currentFill
sByType.cell(row = 1, column = typeDict["Review"]).fill = currentFill
sByType.cell(row = 1, column = typeDict["Review"]+1).fill = currentFill

# Red border around Overdue, Complete, Review, Review+1
sByType.cell(row = 1, column = typeDict["Overdue"]-1).border = Border(top=redBorder, left=redBorder, bottom=redBorder)
sByType.cell(row = 1, column = typeDict["Overdue"]).border = Border(top=redBorder, bottom=redBorder)
sByType.cell(row = 1, column = typeDict["Complete"]).border = Border(top=redBorder, bottom=redBorder)
sByType.cell(row = 1, column = typeDict["Review"]).border = Border(top=redBorder, bottom=redBorder)
sByType.cell(row = 1, column = typeDict["Review"]+1).border = Border(top=redBorder, right=redBorder, bottom=redBorder)

sByType.cell(row = sByType.max_row, column = typeDict["Overdue"]-1).border = Border(top=redBorder, left=redBorder, bottom=redBorder)
sByType.cell(row = sByType.max_row, column = typeDict["Overdue"]).border = Border(top=redBorder, bottom=redBorder)
sByType.cell(row = sByType.max_row, column = typeDict["Complete"]).border = Border(top=redBorder, bottom=redBorder)
sByType.cell(row = sByType.max_row, column = typeDict["Review"]).border = Border(top=redBorder, bottom=redBorder)
sByType.cell(row = sByType.max_row, column = typeDict["Review"]+1).border = Border(top=redBorder, right=redBorder, bottom=redBorder)


# *****************************************************************
# *********************** ByClient Tab ****************************
# *****************************************************************

sByClient = wbReport["ByClient"]

clientDict = {}
get_column_names_and_index(sByClient, clientDict)

# Freeze and left-align top row
sByClient.freeze_panes = "A2"

# Format currency cols to ()
for r in range(2, sByClient.max_row+1):
    for c in range(2, sByClient.max_column+1):
        sByClient.cell(row=r, column=c).number_format = '"$"#,##0_);[Red]("$"#,##0)' # What a gross formatting string...

# Format Column Widths
for col in range(2, sByClient.max_column):
    sByClient.column_dimensions[get_column_letter(col)].width = 12
sByClient.column_dimensions[get_column_letter(1)].width = 40
sByClient.column_dimensions[get_column_letter(sByClient.max_column)].width = 14

# Color Total Rows and Cols blue and make bold
for c in range(1, sByClient.max_column):
    sByClient.cell(row = sByClient.max_row, column = c).fill = blueFill
    sByClient.cell(row = sByClient.max_row, column = c).font = boldFont
for r in range(2, sByClient.max_row+1):
    sByClient.cell(row = r, column = clientDict["Total"]).fill = blueFill
    sByClient.cell(row = r, column = clientDict["Total"]).font = boldFont 
    sByClient.cell(row = r, column = 1).alignment = leftAlign

# Color Overdue, Complete, Review, Review+1 #E6B8B7
sByClient.cell(row = 1, column = clientDict["Overdue"]-1).fill = currentFill
sByClient.cell(row = 1, column = clientDict["Overdue"]).fill = currentFill
sByClient.cell(row = 1, column = clientDict["Complete"]).fill = currentFill
sByClient.cell(row = 1, column = clientDict["Review"]).fill = currentFill
sByClient.cell(row = 1, column = clientDict["Review"]+1).fill = currentFill

# Red border around Overdue, Complete, Review, Review+1
sByClient.cell(row = 1, column = clientDict["Overdue"]-1).border = Border(top=redBorder, left=redBorder, bottom=redBorder)
sByClient.cell(row = 1, column = clientDict["Overdue"]).border = Border(top=redBorder, bottom=redBorder)
sByClient.cell(row = 1, column = clientDict["Complete"]).border = Border(top=redBorder, bottom=redBorder)
sByClient.cell(row = 1, column = clientDict["Review"]).border = Border(top=redBorder, bottom=redBorder)
sByClient.cell(row = 1, column = clientDict["Review"]+1).border = Border(top=redBorder, right=redBorder, bottom=redBorder)

sByClient.cell(row = sByClient.max_row, column = clientDict["Overdue"]-1).border = Border(top=redBorder, left=redBorder, bottom=redBorder)
sByClient.cell(row = sByClient.max_row, column = clientDict["Overdue"]).border = Border(top=redBorder, bottom=redBorder)
sByClient.cell(row = sByClient.max_row, column = clientDict["Complete"]).border = Border(top=redBorder, bottom=redBorder)
sByClient.cell(row = sByClient.max_row, column = clientDict["Review"]).border = Border(top=redBorder, bottom=redBorder)
sByClient.cell(row = sByClient.max_row, column = clientDict["Review"]+1).border = Border(top=redBorder, right=redBorder, bottom=redBorder)

# *****************************************************************
# ************************ PM Tab *********************************
# *****************************************************************

sPM = wbReport["PM"]

PMDict = {}
get_column_names_and_index(sPM, PMDict)

# Freeze and left-align top row
sPM.freeze_panes = "A2"

# Format currency cols to ()
for r in range(2, sPM.max_row+1):
    for c in range(2, sPM.max_column+1):
        sPM.cell(row=r, column=c).number_format = '"$"#,##0_);[Red]("$"#,##0)'

# Format Column Widths
for col in range(2, sPM.max_column):
    sPM.column_dimensions[get_column_letter(col)].width = 12
sPM.column_dimensions[get_column_letter(1)].width = 25
sPM.column_dimensions[get_column_letter(sPM.max_column)].width = 14

# Color Total Rows and Cols blue and make bold
for c in range(1, sPM.max_column):
    sPM.cell(row = sPM.max_row, column = c).fill = blueFill
    sPM.cell(row = sPM.max_row, column = c).font = boldFont
for r in range(2, sPM.max_row+1):
    sPM.cell(row = r, column = PMDict["Total"]).fill = blueFill
    sPM.cell(row = r, column = PMDict["Total"]).font = boldFont
    sPM.cell(row = r, column = 1).alignment = leftAlign


# Color Overdue, Complete, Review, Review+1 #E6B8B7
sPM.cell(row = 1, column = PMDict["Overdue"]-1).fill = currentFill
sPM.cell(row = 1, column = PMDict["Overdue"]).fill = currentFill
sPM.cell(row = 1, column = PMDict["Complete"]).fill = currentFill
sPM.cell(row = 1, column = PMDict["Review"]).fill = currentFill
sPM.cell(row = 1, column = PMDict["Review"]+1).fill = currentFill

# Red border around Overdue, Complete, Review, Review+1
sPM.cell(row = 1, column = PMDict["Overdue"]-1).border = Border(top=redBorder, left=redBorder, bottom=redBorder)
sPM.cell(row = 1, column = PMDict["Overdue"]).border = Border(top=redBorder, bottom=redBorder)
sPM.cell(row = 1, column = PMDict["Complete"]).border = Border(top=redBorder, bottom=redBorder)
sPM.cell(row = 1, column = PMDict["Review"]).border = Border(top=redBorder, bottom=redBorder)
sPM.cell(row = 1, column = PMDict["Review"]+1).border = Border(top=redBorder, right=redBorder, bottom=redBorder)

sPM.cell(row = sPM.max_row, column = PMDict["Overdue"]-1).border = Border(top=redBorder, left=redBorder, bottom=redBorder)
sPM.cell(row = sPM.max_row, column = PMDict["Overdue"]).border = Border(top=redBorder, bottom=redBorder)
sPM.cell(row = sPM.max_row, column = PMDict["Complete"]).border = Border(top=redBorder, bottom=redBorder)
sPM.cell(row = sPM.max_row, column = PMDict["Review"]).border = Border(top=redBorder, bottom=redBorder)
sPM.cell(row = sPM.max_row, column = PMDict["Review"]+1).border = Border(top=redBorder, right=redBorder, bottom=redBorder)

# *****************************************************************
# *********************** NY-PM Tab *******************************
# *****************************************************************

NYsPM = wbReport["NY-PM"]

NYPMDict = {}
get_column_names_and_index(NYsPM, NYPMDict)

# Freeze and left-align top row
NYsPM.freeze_panes = "A2"

# Format currency cols to ()
for r in range(2, NYsPM.max_row+1):
    for c in range(2, NYsPM.max_column+1):
        NYsPM.cell(row=r, column=c).number_format = '"$"#,##0_);[Red]("$"#,##0)'

# Format Column Widths
for col in range(2, NYsPM.max_column):
    NYsPM.column_dimensions[get_column_letter(col)].width = 12
NYsPM.column_dimensions[get_column_letter(1)].width = 25
NYsPM.column_dimensions[get_column_letter(NYsPM.max_column)].width = 14

# Color Total Rows and Cols blue and make bold
for c in range(1, NYsPM.max_column):
    NYsPM.cell(row = NYsPM.max_row, column = c).fill = blueFill
    NYsPM.cell(row = NYsPM.max_row, column = c).font = boldFont
for r in range(2, NYsPM.max_row+1):
    NYsPM.cell(row = r, column = NYPMDict["Total"]).fill = blueFill
    NYsPM.cell(row = r, column = NYPMDict["Total"]).font = boldFont
    NYsPM.cell(row = r, column = 1).alignment = leftAlign


# *****************************************************************
# ******************** Summary Tab ********************************
# *****************************************************************

sSummary = wbReport["Summary"]

summaryDict = {}
get_column_names_and_index(sSummary, summaryDict)

# Freeze and left-align top row
sSummary.freeze_panes = "A2"

# Format currency cols to ()
for r in range(2, sSummary.max_row+1):
    for c in range(3, sSummary.max_column+1):
        sSummary.cell(row=r, column=c).number_format = '"$"#,##0_);[Red]("$"#,##0)'
for r in range(2, sSummary.max_row):
        sSummary.cell(row = r, column = 1).alignment = leftAlign
        sSummary.cell(row = r, column = 2).alignment = leftAlign

# Format Column Widths
for col in range(2, sSummary.max_column):
    sSummary.column_dimensions[get_column_letter(col)].width = 11.6
sSummary.column_dimensions[get_column_letter(1)].width = 14.5
sSummary.column_dimensions[get_column_letter(2)].width = 29
sSummary.column_dimensions[get_column_letter(sSummary.max_column)].width = 13.5

# Color Total Rows and Cols blue and make bold
for r in range(2, sSummary.max_row+1):
    sSummary.cell(row = r, column = summaryDict["Total"]).fill = blueFill
    sSummary.cell(row = r, column = summaryDict["Total"]).font = boldFont
     # IF ClientName (col=2) contains "Total", Format whole row Blue     
    if (re.search("Total", sSummary.cell(row=r, column=2).value)):
        for c in range(2, sSummary.max_column):
            sSummary.cell(row=r, column=c).fill = blueFill 
            sSummary.cell(row=r, column=c).font = boldFont
            sSummary.cell(row = r, column = 1).alignment = leftAlign
            sSummary.cell(row = r, column = 2).alignment = leftAlign
            

# Color Overdue, Complete, Review, Review+1 #E6B8B7
sSummary.cell(row = 1, column = summaryDict["Overdue"]-1).fill = currentFill
sSummary.cell(row = 1, column = summaryDict["Overdue"]).fill = currentFill
sSummary.cell(row = 1, column = summaryDict["Complete"]).fill = currentFill
sSummary.cell(row = 1, column = summaryDict["Review"]).fill = currentFill
sSummary.cell(row = 1, column = summaryDict["Review"]+1).fill = currentFill

# Red border around Overdue, Complete, Review, Review+1
sSummary.cell(row = 1, column = summaryDict["Overdue"]-1).border = Border(top=redBorder, left=redBorder, bottom=redBorder)
sSummary.cell(row = 1, column = summaryDict["Overdue"]).border = Border(top=redBorder, bottom=redBorder)
sSummary.cell(row = 1, column = summaryDict["Complete"]).border = Border(top=redBorder, bottom=redBorder)
sSummary.cell(row = 1, column = summaryDict["Review"]).border = Border(top=redBorder, bottom=redBorder)
sSummary.cell(row = 1, column = summaryDict["Review"]+1).border = Border(top=redBorder, right=redBorder, bottom=redBorder)

sSummary.cell(row = sSummary.max_row, column = summaryDict["Overdue"]-1).border = Border(top=redBorder, left=redBorder, bottom=redBorder)
sSummary.cell(row = sSummary.max_row, column = summaryDict["Overdue"]).border = Border(top=redBorder, bottom=redBorder)
sSummary.cell(row = sSummary.max_row, column = summaryDict["Complete"]).border = Border(top=redBorder, bottom=redBorder)
sSummary.cell(row = sSummary.max_row, column = summaryDict["Review"]).border = Border(top=redBorder, bottom=redBorder)
sSummary.cell(row = sSummary.max_row, column = summaryDict["Review"]+1).border = Border(top=redBorder, right=redBorder, bottom=redBorder)




# Save Workbook
wbReport.save(outputFName)
print("Formatting completed. Saved to " + outputFName + "\n")