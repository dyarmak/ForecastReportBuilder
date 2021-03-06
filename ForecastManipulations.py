#************************************************************************
# This script performs the excel manipulations on the qry_Forecast data
#************************************************************************

# Currently broken
# Need to add 'Type' Column in correct loaction

import os
import datetime
import time
import openpyxl
from openpyxl.styles import Alignment, PatternFill
from myxlutils import get_column_names_and_index
from excelFNames import forecastFName, logMe

startTimer = time.time()

# ------------- Load ---------------
wbFore = openpyxl.load_workbook(forecastFName)
sFore = wbFore.active

# ----------- BEGIN Find Column Indexes -------------------
print("Loading Forecast column names")
#Create an empty dictionary
forecastDict = {}
# fill the dictionary with:
# Key = column names
# Value = column index
get_column_names_and_index(sFore, forecastDict)

# -------------- BEGIN Formatting -----------

# ----------- BEGIN Find and Delete ------------
# Delete WHERE SubProjectStatus == "planning" && IncludeinBudget == False && InvoicedAmount == 0)

# Need to add invAmount test here as well?

print("Removing rows WHERE Status == Planning AND inBudget == False\n")

if(logMe == 1):
    deleteLog = open("deleteLog.txt", "w+")

# find and Delete WHERE Status==Planning and inBudget==False
# Start from the bottom, because :func delete_rows() shifts the rows below it up after it deletes a row, and would therefore skip a row
for r in range(sFore.max_row+1, 2, -1): 
    if(sFore.cell(row=r, column=forecastDict["SubProjectStatus"]).value == "Planning" and sFore.cell(row=r, column=forecastDict["IncludeinBudget"]).value == False):
        if(logMe == 1):
            #Log the deleted rows
            logString = "deleted subID: " + str(sFore.cell(row=r, column=1).value) + ". Status == Planning and IncludeinBudget == False\n"
            deleteLog.write(logString)
        # Delete this row
        sFore.delete_rows(r, 1)

# Delete rows WHERE SubProjectStatus == "Cancel Requested"
    if(sFore.cell(row=r, column=forecastDict["SubProjectStatus"]).value == "Cancel Requested"):
        if(logMe == 1):
            #Log the deleted rows
            logString = "deleted subID: " + str(sFore.cell(row=r, column=1).value) + " Cancel Requested\n"
            deleteLog.write(logString)
        # Delete this row
        sFore.delete_rows(r, 1)
if(logMe == 1):
    print("Deleted rows logged in deleteLog.txt")
    deleteLog.close()
# ---------- END Find and Delete ------------------


# ------ Remove 0 values from Quoted and set to None -----
for r in range(2, sFore.max_row+1):
    if (sFore.cell(row=r, column=forecastDict["Quoted"]).value == 0): # Quoted = 0 might be due to a quote haviung been started, but not finished
        sFore.cell(row=r, column=forecastDict["Quoted"]).value = None


# --------------Log subProjectIDs with no Forecast, Budget, Quote, or originalForecast and SubprojectStatus != "Complete"
if(logMe == 1):
    print("Logging entries WHERE\nForecast, Budget, Quoted and OriginalForecast are all blank \nAND SubProjectStatus != Complete\n")
    noDollarValues = open("noDollarValuesLog.txt", "w+")
    for r in range(2, sFore.max_row+1):
        if(sFore.cell(row=r, column=forecastDict["Forecast"]).value == None and 
            sFore.cell(row=r, column=forecastDict["Budget"]).value == None and
            sFore.cell(row=r, column=forecastDict["Quoted"]).value == None and
            sFore.cell(row=r, column=forecastDict["OriginalForecast"]).value == None and
            sFore.cell(row=r, column=forecastDict["SubProjectStatus"]).value != "Complete"):
        
            logString = "SubProjectID: " + str(sFore.cell(row=r, column=forecastDict["SubProjectID"]).value) + " has no Dollar values\n"
            noDollarValues.write(logString)
        

    noDollarValues.close()
    print("Created noDollarValues.txt\n")


# ---------- BEGIN Update Forecast Amount -------
print("Updating forecasted amounts\n")

if(logMe == 1):
    forecastLog = open("forecastLog.txt", "w+")

# INSERT Interim invoicing logic here? or Below?

for r in range(2, sFore.max_row+1):
    # IF InvoiceDateSent != None, Forecast GETS None
    # cell formatting does not carry over when amalgamating the data in the next step. 
    if(sFore.cell(row=r, column=forecastDict["InvoiceDateSent"]).value != None): # May need to test for == 2018 or == 2019 instead
        if(logMe == 1):
            logString = "subID " + str(sFore.cell(row=r, column=1).value) + " has an Invoice\n"
            forecastLog.write(logString)
        sFore.cell(row=r, column=forecastDict["Forecast"]).value = None

    # There is no InvoiceDateSent. 
    # IF Quoted !=None, Forecast GETS Quoted
    elif(sFore.cell(row=r, column=forecastDict["Quoted"]).value != None):
        if(logMe == 1):
            logString = "subID " + str(sFore.cell(row=r, column=1).value) + " gets Quoted: "  + str(sFore.cell(row=r, column=forecastDict["Quoted"]).value) + "\n"
            forecastLog.write(logString)
        sFore.cell(row=r, column=forecastDict["Forecast"]).value = sFore.cell(row=r, column=forecastDict["Quoted"]).value

    # There is no InvoiceDateSent AND no Quoted
    # IF OriginalForecast !=None, Forecast GETS OriginalForecast
    elif(sFore.cell(row=r, column=forecastDict["OriginalForecast"]).value is not None):
        if(logMe == 1):
            logString = "subID " + str(sFore.cell(row=r, column=1).value) + " gets OriginalForecast:" + str(sFore.cell(row=r, column=forecastDict["OriginalForecast"]).value) + "\n"
            forecastLog.write(logString)
        sFore.cell(row=r, column=forecastDict["Forecast"]).value = sFore.cell(row=r, column=forecastDict["OriginalForecast"]).value
    
    # There is no InvoiceDateSent AND no Quoted AND no OriginatlForecast
    # IF there is a budget, Forecast GETS budget
    else:
        if(logMe == 1):
            logString = "subID " + str(sFore.cell(row=r, column=1).value) + " gets Budget: " + str(sFore.cell(row=r, column=forecastDict["Budget"]).value) + "\n"
            forecastLog.write(logString)
        sFore.cell(row=r, column=forecastDict["Forecast"]).value = sFore.cell(row=r, column=forecastDict["Budget"]).value

# Having Quoted be last means its could overwrite... Should maybe change the order of this, since Quoted should be the default...?

if(logMe == 1):
    forecastLog.close()
    print("Created forecastLog.txt")
# --------- END Update Forecast Amount ---------


# -------------- Update Due Column ---------------

# Need todays date for date compare
todayDate = datetime.date.today()

if(logMe == 1):
    overdueLog = open("overdueLog.txt", "w+")

for r in range(2, sFore.max_row+1):
# IF subprojectStatus == "Review" or == "Complete", Set Due column to SubProjectStatus
    if(sFore.cell(row=r, column=forecastDict["SubProjectStatus"]).value == "Review" or sFore.cell(row=r, column=forecastDict["SubProjectStatus"]).value == "Complete"):
        sFore.cell(row=r, column=forecastDict["Due"]).value = sFore.cell(row=r, column=forecastDict["SubProjectStatus"]).value

    # IF the ProjectStatus == "In Progess" or == "Planning" AND DueDate => today, set Due to "Overdue"
    if(sFore.cell(row=r, column=forecastDict["SubProjectStatus"]).value == "In Progress" or sFore.cell(row=r, column=forecastDict["SubProjectStatus"]).value == "Planning"):
        dueDate = sFore.cell(row=r, column=forecastDict["Due Date"]).value.date() # read date without time from excel value
        if(dueDate<todayDate):
            if(logMe == 1):
                logString = "subID " + str(sFore.cell(row=r, column=1).value) + " is overdue\n"
                overdueLog.write(logString)
            sFore.cell(row=r, column=forecastDict["Due"]).value = "Overdue"

# CORNER CASE
# What about values In Progress or Planning that are not overdue?
if(logMe == 1):
    overdueLog.close()
    print("Created overdueLog.txt")
# -----------END Update Due Column ---------------

# ------------ SAVE Output Excel File -------------

wbFore.save(forecastFName)
print("Manipulated Forecast query saved as " + forecastFName)
wbFore.close()

# -------------- Timer -----------------------
endTimer = time.time()
forecastTime = endTimer-startTimer
print("Execution on Forecast took: " + str(forecastTime) + "seconds to execute\n")

# input("Press enter to exit")

