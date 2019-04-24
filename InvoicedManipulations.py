#************************************************************************
# This script performs the excel manipulations on the qry_Invoiced data
#************************************************************************

# Currently broken
# Need to add 'Type' Column in correct loaction

import os
import datetime
import time
import xlrd
from openpyxl import Workbook
import openpyxl
from myxlutils import save_and_reopen, format_date_rows, get_column_names_and_index
from excelFNames import invoicedFName, logMe

startTimer = time.time()

# ------------- Load ---------------
wbInvo = openpyxl.load_workbook(invoicedFName)
sInvo = wbInvo.active 

# ----------- BEGIN Find Column Indexes -------------------
# Should have a dictionary that stores the column names and index position

print("Loading Invoiced column names")
# Create an empty dictionary
invoicedDict = {}
# fill the dictionary with:
# Key = column names
# Value = column index
get_column_names_and_index(sInvo, invoicedDict)


# ------------- Deferred Revenues ---------------
# IF InvoiceDateSent == 2018   Due Gets "Def-1"
if(logMe == 1):
    invoicedDeferredLog = open("invoicedDeferredLog.txt", "w+") 

for r in range(2, sInvo.max_row+1):
    if(sInvo.cell(row=r, column=invoicedDict["InvoiceDateSent"]).value != None and sInvo.cell(row=r, column=invoicedDict["InvoiceDateSent"]).value.date().year == 2018):
        sInvo.cell(row=r, column=invoicedDict["Due"]).value = "Def-1"
        

# ------------ SAVE Output Excel File -------------

wbInvo.save(invoicedFName)
print("Manipulated Invoiced query saved as " + invoicedFName)
wbInvo.close()

# -------------- Timer -----------------------
endTimer = time.time()
invoicedTime = endTimer-startTimer
print("Execution on Invoiced took: " + str(invoicedTime) + "seconds to execute\n")
