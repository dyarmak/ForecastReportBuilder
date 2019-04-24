import os
import datetime
import openpyxl
from myxlutils import get_column_names_and_index
from excelFNames import forecastFName, invoicedFName, creditFName


# ------------- Load Workbooks ---------------
wbFore = openpyxl.load_workbook(forecastFName)
sFore = wbFore.active

wbInvo = openpyxl.load_workbook(invoicedFName)
sInvo = wbInvo.active

wbCred = openpyxl.load_workbook(creditFName)
sCred = wbCred.active

# Create empty dictionaries
forecastHeading = {}
invoicedHeading = {}
creditsHeading = {}

# Get column names
get_column_names_and_index(sFore, forecastHeading)
get_column_names_and_index(sInvo, invoicedHeading)
get_column_names_and_index(sCred, creditsHeading)


for r in range(2, sFore.max_row+1):
    MMYYYY = sFore.cell(row=r, column=forecastHeading["Due Date"]).value.date()
    MMYYYY = MMYYYY.strftime("%m-%Y")
    sFore.cell(row=r, column=forecastHeading["Due"]).value = MMYYYY
wbFore.save(forecastFName)

for r in range(2, sInvo.max_row+1):
    MMYYYY = sInvo.cell(row=r, column=invoicedHeading["Due Date"]).value.date()
    MMYYYY = MMYYYY.strftime("%m-%Y")
    sInvo.cell(row=r, column=invoicedHeading["Due"]).value = "Act" + MMYYYY
wbInvo.save(invoicedFName)

for r in range(2, sCred.max_row+1):
    MMYYYY = sCred.cell(row=r, column=creditsHeading["Due Date"]).value.date()
    MMYYYY = MMYYYY.strftime("%m-%Y")
    sCred.cell(row=r, column=creditsHeading["Due"]).value = "Act" + MMYYYY
wbCred.save(creditFName)

print("Values of Due column set\n")