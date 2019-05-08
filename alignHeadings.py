import os
import openpyxl
from myxlutils import get_column_names_and_index
from excelFNames import forecastFName, invoicedFName, creditFName

# -------------------------- FORECAST ----------------------------------------
# Forecast is missing "Type" and "Due"
wbFore = openpyxl.load_workbook(forecastFName)
sFore = wbFore.active
forecastHeading = {}
get_column_names_and_index(sFore, forecastHeading)
# Insert "Type" after SubProjectTypeName
sFore.insert_cols(forecastHeading["SubProjectTypeName"]+1)
sFore.cell(row=1, column= forecastHeading["SubProjectTypeName"]+1).value = "Type"
# Update indexes after move
get_column_names_and_index(sFore, forecastHeading)
# Insert "Due" after "Due Date"
sFore.insert_cols(forecastHeading["Due Date"]+1)
sFore.cell(row=1, column= forecastHeading["Due Date"]+1).value = "Due"


# -------------------------- INVOICED ----------------------------------------
# Invoiced is missing "Type" and "Due"
wbInvo = openpyxl.load_workbook(invoicedFName)
sInvo = wbInvo.active
invoicedHeading = {}
get_column_names_and_index(sInvo, invoicedHeading)
# Insert "Type" after SubProjectTypeName
sInvo.insert_cols(invoicedHeading["SubProjectTypeName"]+1)
sInvo.cell(row=1, column= invoicedHeading["SubProjectTypeName"]+1).value = "Type"
# Update indexes after move
get_column_names_and_index(sInvo, invoicedHeading)
# Insert "Due" after "Due Date"
sInvo.insert_cols(invoicedHeading["Due Date"]+1)
sInvo.cell(row=1, column= invoicedHeading["Due Date"]+1).value = "Due"
get_column_names_and_index(sInvo, invoicedHeading)
# This is NOT in Christy's process, but in order to make the data actually line up we need to...
# Insert 3 rows after "OriginalDueDate"
for x in range(0,3):
    sInvo.insert_cols(invoicedHeading["OriginalDueDate"]+1)
# Update indexes after move
get_column_names_and_index(sInvo, invoicedHeading)


# -------------------------- CREDITS ----------------------------------------
wbCred = openpyxl.load_workbook(creditFName)
sCred = wbCred.active
creditsHeading = {}
get_column_names_and_index(sCred, creditsHeading)
# Insert "Type" after SubProjectTypeName
sCred.insert_cols(creditsHeading["SubProjectTypeName"]+1)
sCred.cell(row=1, column= creditsHeading["SubProjectTypeName"]+1).value = "Type"
# Update indexes after move
get_column_names_and_index(sCred, creditsHeading)
# Insert "Due" after "Due Date"
sCred.insert_cols(creditsHeading["Due Date"]+1)
sCred.cell(row=1, column= creditsHeading["Due Date"]+1).value = "Due"
# Update indexes after move
get_column_names_and_index(sCred, creditsHeading)


# Save all the Workbooks
wbFore.save(forecastFName)
wbFore.close()

wbInvo.save(invoicedFName)
wbInvo.close()

wbCred.save(creditFName)
wbCred.close()

print("Headings aligned\n")