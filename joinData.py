import os
import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from myxlutils import format_date_rows, get_column_names_and_index
from excelFNames import forecastFName, invoicedFName, creditFName, combinedFName


wbCombined = openpyxl.Workbook()
sCombined = wbCombined.active
sCombined.title = "Detail"

wbFore = openpyxl.load_workbook(forecastFName)
sFore = wbFore.active

wbInvo = openpyxl.load_workbook(invoicedFName)
sInvo = wbInvo.active

wbCred = openpyxl.load_workbook(creditFName)
sCred = wbCred.active

# -------------------- Copy Forecast -> Combined ------------------------------------- 
foreRowCount = 0

# We include the column names from Forecast.xlsx, not for other 2 sheets
for r in range(1, sFore.max_row+1):
        for c in range(1, sFore.max_column+1):
                sCombined.cell(row=r,column=c).value = sFore.cell(row=r, column=c).value
        foreRowCount+=1

wbCombined.save(combinedFName)
wbCombined.close()
wbCombined = openpyxl.load_workbook(combinedFName)
sCombined = wbCombined.active

combinedDict = {}
get_column_names_and_index(sCombined,combinedDict)
format_date_rows(sCombined, combinedDict, "mm-dd-yy", "Due Date", "InvoiceDateSent", "OriginalDueDate")

yellowFill = PatternFill(patternType="solid", fgColor="FFFF00" )

for r in range(2, sCombined.max_row+1):
        # IF InvoiceDateSent != None, Forecast GOT None But formatting doesn't copy over. 
        # We NOW can fill cell background with yellow
        # HAVE to do this step AFTER copying sFore, and BEFORE copying anything else...
        if(sCombined.cell(row=r, column=combinedDict["InvoiceDateSent"]).value != None): # May need to test for == 2018 or == 2019 instead
                sCombined.cell(row=r, column=combinedDict["Forecast"]).fill = yellowFill 

# We want to start
combinedStartPoint = foreRowCount -1
invoRowCount = 0        
for r in range(2, sInvo.max_row+1):
        for c in range(1, sInvo.max_column+1):
                sCombined.cell(row=r+combinedStartPoint,column=c).value = sInvo.cell(row=r, column=c).value
        invoRowCount+=1
        

wbCombined.save(combinedFName)
combinedStartPoint = foreRowCount + invoRowCount -1
credRowCount = 0
for r in range(2, sCred.max_row+1):
        for c in range(1, sCred.max_column+1):
                sCombined.cell(row=r+combinedStartPoint,column=c).value = sCred.cell(row=r, column=c).value
        credRowCount+=1

wbCombined.save(combinedFName)

# Save and re-load workbook
wbCombined.save(combinedFName)
wbCombined.close()
wbCombined = openpyxl.load_workbook(combinedFName)
sCombined = wbCombined.active
# Re-load column names into a new dict

# ----------------- Formatting -------------------------
sCombined.freeze_panes = "A2"

alignment = Alignment(horizontal="left", vertical="top")
for x in range(1,(sCombined.max_column+1)):
    sCombined.cell(row=1, column=x).alignment = alignment

format_date_rows(sCombined, combinedDict, "mm-dd-yy", "Due Date", "InvoiceDateSent", "OriginalDueDate")

sCombined.column_dimensions[get_column_letter(combinedDict['Due Date'])].width = 15.2
sCombined.column_dimensions[get_column_letter(combinedDict['Due'])].width = 15.2
sCombined.column_dimensions[get_column_letter(combinedDict['InvoiceDateSent'])].width = 15.2
sCombined.column_dimensions[get_column_letter(combinedDict['OriginalDueDate'])].width = 15.2

# Save and close
wbCombined.save(combinedFName)
wbCombined.close()

print("Forecast Row Count: " + str(foreRowCount))
print("Invoiced Row Count: " + str(invoRowCount))
print("Credits Row Count: " + str(credRowCount))
totalRows = foreRowCount + invoRowCount + credRowCount
print("Total Rows Joined: " + str(totalRows))
print("Data from 3 queries combined\n")